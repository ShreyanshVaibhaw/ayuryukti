"""Analytics report generation and export utilities."""

from __future__ import annotations

from datetime import datetime, timezone
from pathlib import Path
from typing import Dict

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle


def generate_analytics_pdf(analytics_data: Dict, output_path: str):
    """Generate treatment effectiveness analytics PDF."""
    out = Path(output_path)
    out.parent.mkdir(parents=True, exist_ok=True)

    effectiveness: pd.DataFrame = analytics_data.get("effectiveness", pd.DataFrame())
    response: Dict = analytics_data.get("prakriti_response", {})

    styles = getSampleStyleSheet()
    doc = SimpleDocTemplate(str(out), pagesize=A4)
    story = []

    story.append(Paragraph("<b>AyurYukti YuktiShaala — Treatment Analytics</b>", styles["Title"]))
    story.append(Paragraph(f"Generated: {datetime.now(timezone.utc).isoformat()} UTC", styles["BodyText"]))
    story.append(Spacer(1, 10))

    story.append(Paragraph("<b>Condition-wise Effectiveness</b>", styles["Heading3"]))
    if effectiveness.empty:
        story.append(Paragraph("No outcome data available.", styles["BodyText"]))
    else:
        rows = [["Formulation", "Prakriti", "N", "Success Rate", "95% CI"]]
        for _, row in effectiveness.head(20).iterrows():
            rows.append(
                [
                    str(row.get("formulation", "")),
                    str(row.get("prakriti", "")),
                    str(int(row.get("n_patients", 0))),
                    f"{float(row.get('success_rate', 0.0)) * 100:.1f}%",
                    f"{float(row.get('ci_low', 0.0)):.2f}-{float(row.get('ci_high', 0.0)):.2f}",
                ]
            )
        table = Table(rows, colWidths=[120, 90, 40, 90, 90])
        table.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#f2f5f9")),
                    ("GRID", (0, 0), (-1, -1), 0.4, colors.grey),
                    ("FONTSIZE", (0, 0), (-1, -1), 8),
                ]
            )
        )
        story.append(table)

    story.append(Spacer(1, 10))
    story.append(Paragraph("<b>Prakriti Response Analysis</b>", styles["Heading3"]))
    if not response:
        story.append(Paragraph("No response matrix available.", styles["BodyText"]))
    else:
        for prakriti, formulations in response.items():
            top = sorted(formulations.items(), key=lambda x: x[1], reverse=True)[:3]
            top_txt = ", ".join([f"{name} ({rate * 100:.1f}%)" for name, rate in top])
            story.append(Paragraph(f"- {prakriti}: {top_txt}", styles["BodyText"]))

    doc.build(story)


def export_analytics_excel(analytics_data: Dict, output_path: str):
    """Export summary, prakriti analysis, and raw outcomes to Excel."""
    out = Path(output_path)
    out.parent.mkdir(parents=True, exist_ok=True)

    effectiveness: pd.DataFrame = analytics_data.get("effectiveness", pd.DataFrame())
    response: Dict = analytics_data.get("prakriti_response", {})
    raw: pd.DataFrame = analytics_data.get("raw_outcomes", pd.DataFrame())

    wb = Workbook()
    ws_summary = wb.active
    ws_summary.title = "Summary"

    header_fill = PatternFill(start_color="FFEEECE1", end_color="FFEEECE1", fill_type="solid")
    header_font = Font(bold=True)

    summary_headers = ["Formulation", "Prakriti", "N", "Success Rate", "CI Low", "CI High"]
    ws_summary.append(summary_headers)
    for cell in ws_summary[1]:
        cell.fill = header_fill
        cell.font = header_font

    for _, row in effectiveness.iterrows():
        ws_summary.append(
            [
                row.get("formulation", ""),
                row.get("prakriti", ""),
                int(row.get("n_patients", 0)),
                float(row.get("success_rate", 0.0)),
                float(row.get("ci_low", 0.0)),
                float(row.get("ci_high", 0.0)),
            ]
        )

    ws_pr = wb.create_sheet("Prakriti Analysis")
    ws_pr.append(["Prakriti", "Formulation", "Success Rate"])
    for cell in ws_pr[1]:
        cell.fill = header_fill
        cell.font = header_font

    for prakriti, formulations in response.items():
        for name, rate in formulations.items():
            ws_pr.append([prakriti, name, float(rate)])

    ws_raw = wb.create_sheet("Raw Outcome Data")
    if raw.empty:
        ws_raw.append(["No raw data available"])
    else:
        ws_raw.append(list(raw.columns))
        for cell in ws_raw[1]:
            cell.fill = header_fill
            cell.font = header_font
        for row in raw.itertuples(index=False):
            ws_raw.append(list(row))

    for ws in [ws_summary, ws_pr, ws_raw]:
        ws.auto_filter.ref = ws.dimensions
        for col in ws.columns:
            width = max(len(str(c.value)) if c.value is not None else 0 for c in col) + 2
            ws.column_dimensions[col[0].column_letter].width = min(40, width)

    wb.save(str(out))
